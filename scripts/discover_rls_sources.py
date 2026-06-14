"""
Power BI RLS Source Discovery
==============================
Scans every semantic model in a Power BI workspace and writes
rls_sources_manifest.yaml — one entry per (dataset × RLS role) that uses
USERPRINCIPALNAME(), recording the source file path, format, and UPN column.

The manifest is the configuration input for downstream RLS validation tests.

Usage
-----
  python scripts/discover_rls_sources.py              # interactive
  python scripts/discover_rls_sources.py --all        # scan all workspaces
  python scripts/discover_rls_sources.py --workspace "Finance Analytics"
  python scripts/discover_rls_sources.py --no-xmla    # Tier 1 (REST) only
  python scripts/discover_rls_sources.py --verbose    # print all rows
  python scripts/discover_rls_sources.py --debug      # maximum diagnostic output

Auth
----
Reuses playwright/.auth/msal-device-token-cache.json written by `npm run setup`.
If absent or expired, initiates device-flow and writes a fresh cache.

XMLA (Tier 2)
-------------
Install pyadomd (Windows) for full DAX role + M expression extraction.
Without it the script falls back to Tier 1 (REST datasource API) which gives
file paths but not UPN column names.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
import yaml

# ── optional deps ──────────────────────────────────────────────────────────────

try:
    import msal
    _MSAL_OK = True
except ImportError:
    _MSAL_OK = False

try:
    import openpyxl as _openpyxl
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

try:
    import csv as _csv_mod
    _CSV_OK = True
except ImportError:
    _CSV_OK = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── SSMS 21 ADOMD DLL paths (validated) ───────────────────────────────────────
_DLL_PATHS = {
    "CORE":    r"C:\Program Files\Microsoft SQL Server Management Studio 21\Release\Common7\IDE\Microsoft.AnalysisServices.Core.dll",
    "TABULAR": r"C:\Program Files\Microsoft SQL Server Management Studio 21\Release\Common7\IDE\Microsoft.AnalysisServices.Tabular.dll",
    "ADOMD":   r"C:\Program Files\Microsoft SQL Server Management Studio 21\Release\Common7\IDE\Microsoft.AnalysisServices.AdomdClient.dll",
}

_XMLA_LIB_OK = False
_XMLA_DLL_MISSING: list[str] = []
try:
    import clr as _clr  # pythonnet — required by pyadomd
    for _label, _dll in _DLL_PATHS.items():
        if os.path.isfile(_dll):
            _clr.AddReference(_dll)
        else:
            _XMLA_DLL_MISSING.append(f"{_label}: {_dll}")
    if not _XMLA_DLL_MISSING:
        import pyadomd  # noqa: F401
        from pyadomd import Pyadomd  # noqa: F401
        _XMLA_LIB_OK = True
except Exception:
    _XMLA_LIB_OK = False

# ── constants ──────────────────────────────────────────────────────────────────

LEGACY_CLIENT_ID = "d3590ed6-52b3-4102-aeff-aad2292ab01c"
AUTHORITY_BASE   = "https://login.microsoftonline.com"
SCOPES           = ["https://analysis.windows.net/powerbi/api/.default"]
API_BASE         = "https://api.powerbi.com"

REPO_ROOT        = Path(__file__).resolve().parent.parent
AUTH_DIR         = REPO_ROOT / "playwright" / ".auth"
TOKEN_CACHE_PATH = AUTH_DIR / "msal-device-token-cache.json"
AUTH_META_PATH   = AUTH_DIR / "auth-meta.json"
RLS_DIR          = REPO_ROOT / "rls_discovery"
MANIFEST_PATH    = RLS_DIR / "rls_sources_manifest.yaml"

TOP_N = 20

# Set to True by --debug flag at startup; controls verbose diagnostic output.
_DEBUG: bool = False


def dbg(msg: str) -> None:
    """Print a debug line — only when --debug is active."""
    if _DEBUG:
        print(f"  {dim('[DBG]')} {dim(msg)}")


FILE_SOURCE_TYPES = {
    "SharePointOnline", "SharePoint", "File",
    "Web", "OData", "AzureBlobs",
    # On-premises Active Directory / LDAP (RLS table sourced from AD security groups)
    "ActiveDirectory", "Ldap", "Extension",
}

# On-prem AD / LDAP M expression patterns — checked before file patterns
# so that an AD source is never misclassified as "unknown" file format.
_AD_M_PATTERNS = [
    (re.compile(r'ActiveDirectory\.Domains\s*\(\s*"([^"]+)"',  re.IGNORECASE), "active_directory"),
    (re.compile(r'ActiveDirectory\.Groups\s*\(\s*"([^"]+)"',   re.IGNORECASE), "active_directory"),
    (re.compile(r'ActiveDirectory\.Users\s*\(\s*"([^"]+)"',    re.IGNORECASE), "active_directory"),
    (re.compile(r'Ldap\.Contents\s*\(\s*"([^"]+)"',            re.IGNORECASE), "ldap"),
]

# Regex patterns for extracting file paths from M/Power Query expressions
_M_PATH_PATTERNS = [
    re.compile(r'Excel\.Workbook\s*\(\s*File\.Contents\s*\(\s*"([^"]+)"', re.IGNORECASE),
    re.compile(r'Csv\.Document\s*\(\s*File\.Contents\s*\(\s*"([^"]+)"',   re.IGNORECASE),
    re.compile(r'SharePoint\.Files\s*\(\s*"([^"]+)"',                     re.IGNORECASE),
    re.compile(r'SharePoint\.Tables\s*\(\s*"([^"]+)"',                    re.IGNORECASE),
    re.compile(r'Web\.Contents\s*\(\s*"([^"]+)"',                         re.IGNORECASE),
    re.compile(r'File\.Contents\s*\(\s*"([^"]+)"',                        re.IGNORECASE),
]
_M_SHEET_PATTERN  = re.compile(r'\[Item\s*=\s*"([^"]+)"\]', re.IGNORECASE)
_DAX_UPN_PATTERN  = re.compile(
    r'\[([^\]]+)\]\s*=\s*USERPRINCIPALNAME\s*\(\s*\)',
    re.IGNORECASE,
)

# ── colours ────────────────────────────────────────────────────────────────────

def _enable_ansi() -> bool:
    """Enable ANSI escape processing — returns True if colours should be used."""
    if not sys.stdout.isatty():
        return False
    if sys.platform == "win32":
        try:
            # Enable ENABLE_VIRTUAL_TERMINAL_PROCESSING on Windows 10+ console
            import ctypes
            kernel32 = ctypes.windll.kernel32
            handle = kernel32.GetStdHandle(-11)          # STD_OUTPUT_HANDLE
            mode = ctypes.c_ulong(0)
            if kernel32.GetConsoleMode(handle, ctypes.byref(mode)):
                kernel32.SetConsoleMode(handle, mode.value | 0x0004)
        except Exception:
            pass
        # Also try colorama as an extra fallback
        try:
            import colorama
            colorama.init(wrap=True, strip=False)
        except ImportError:
            pass
    return True

_NO_COLOUR = not _enable_ansi()

def _c(code: str, text: str) -> str:
    return text if _NO_COLOUR else f"\x1b[{code}m{text}\x1b[0m"

def bold(s: str)    -> str: return _c("1",  s)
def dim(s: str)     -> str: return _c("2",  s)
def cyan(s: str)    -> str: return _c("36", s)
def green(s: str)   -> str: return _c("32", s)
def yellow(s: str)  -> str: return _c("33", s)
def red(s: str)     -> str: return _c("31", s)
def magenta(s: str) -> str: return _c("35", s)

# ── timestamp helpers ──────────────────────────────────────────────────────────

def ts() -> str:
    n = datetime.now()
    return dim(f"[{n.hour:02d}:{n.minute:02d}:{n.second:02d}]")

def elapsed(start: float) -> str:
    return dim(f"+{time.monotonic() - start:.1f}s")

def now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

# ── interactive helpers ────────────────────────────────────────────────────────

def _print_list(items: list[dict], label: str, total: int | None = None) -> None:
    pool    = total if total is not None else len(items)
    suffix  = (
        dim(f" — showing {len(items)} of {pool}")
        if pool > len(items)
        else dim(f" — {len(items)} total")
    )
    print(f"\n  {bold(label)}{suffix}")
    for i, item in enumerate(items, 1):
        print(f"    {dim(f'[{i:>3}]')}  {item['name']}")

def pick_one(items: list[dict], label: str) -> dict:
    """Interactive single-select with search."""
    sorted_items = sorted(items, key=lambda x: x["name"].lower())
    visible      = sorted_items[:TOP_N]
    _print_list(visible, label, len(sorted_items))

    while True:
        can_expand = len(sorted_items) > len(visible)
        hint       = (
            dim(f"  type to search · Enter to show all {len(sorted_items)} · ")
            if can_expand else dim("  type to search · ")
        )
        raw = input(f"{hint}Enter number (1–{len(visible)}): ").strip()

        if not raw and can_expand:
            visible = sorted_items
            _print_list(visible, label)
            continue

        if raw.lstrip("/"):
            # /keyword search
            if raw.startswith("/"):
                raw = raw[1:]
            if raw and not raw.isdigit():
                q        = raw.lower()
                filtered = [x for x in sorted_items if q in x["name"].lower()]
                if not filtered:
                    print(yellow(f'  No matches for "{raw}" — showing full list.'))
                    visible = sorted_items
                else:
                    visible = filtered
                _print_list(visible, label, len(sorted_items))
                continue

        try:
            idx = int(raw) - 1
            if 0 <= idx < len(visible):
                return visible[idx]
        except ValueError:
            pass

        print(red(f"  Please enter a number between 1 and {len(visible)}."))

# ── auth ───────────────────────────────────────────────────────────────────────

def _read_auth_meta() -> dict:
    if AUTH_META_PATH.exists():
        try:
            return json.loads(AUTH_META_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"environment": "Public"}

def _save_auth_meta(tenant_id: str, environment: str) -> None:
    AUTH_DIR.mkdir(parents=True, exist_ok=True)
    AUTH_META_PATH.write_text(
        json.dumps({"tenantId": tenant_id, "environment": environment}, indent=2) + "\n",
        encoding="utf-8",
    )

def get_access_token() -> str:
    """
    Acquire a Power BI access token.
    Tries silent acquisition from the shared MSAL cache first.
    Falls back to device-flow when the cache is absent or expired.
    """
    if not _MSAL_OK:
        print(red("  ✗  msal package not found.  Run:  pip install -r scripts/requirements_rls.txt"))
        sys.exit(1)

    meta      = _read_auth_meta()
    tenant_id = meta.get("tenantId", "common")
    authority = f"{AUTHORITY_BASE}/{tenant_id}"

    dbg(f"Auth meta: tenantId={tenant_id!r}  environment={meta.get('environment')!r}")
    dbg(f"Token cache path: {TOKEN_CACHE_PATH}")
    dbg(f"Token cache exists: {TOKEN_CACHE_PATH.exists()}")

    cache = msal.SerializableTokenCache()
    if TOKEN_CACHE_PATH.exists():
        cache.deserialize(TOKEN_CACHE_PATH.read_text(encoding="utf-8"))
        dbg("Token cache deserialised.")

    client_id = os.environ.get("CLIENT_ID", LEGACY_CLIENT_ID)
    dbg(f"Client ID: {client_id}")
    dbg(f"Authority: {authority}")

    app = msal.PublicClientApplication(
        client_id  = client_id,
        authority  = authority,
        token_cache= cache,
    )

    accounts = app.get_accounts()
    dbg(f"Cached accounts: {len(accounts)}  →  {[a.get('username') for a in accounts]}")

    if accounts:
        dbg(f"Attempting silent token acquisition for {accounts[0].get('username')!r} …")
        result = app.acquire_token_silent(SCOPES, account=accounts[0])
        if result and "access_token" in result:
            dbg("Silent acquisition succeeded.")
            _flush_cache(cache)
            return result["access_token"]
        dbg(f"Silent acquisition failed: {result.get('error_description') if result else 'no result'}")

    # Device flow
    print(f"\n  {bold(cyan('Sign in required'))}")
    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        raise RuntimeError(f"Device flow initiation failed: {flow}")
    print(f"  {flow['message']}\n")
    dbg(f"Device flow expires_in: {flow.get('expires_in')}s")
    result = app.acquire_token_by_device_flow(flow)

    if "access_token" not in result:
        raise RuntimeError(f"Authentication failed: {result.get('error_description', result)}")

    dbg("Device flow token acquired.")
    _flush_cache(cache)
    acquired_tenant = result.get("id_token_claims", {}).get("tid", tenant_id)
    dbg(f"Tenant from token: {acquired_tenant!r}")
    _save_auth_meta(acquired_tenant, meta.get("environment", "Public"))
    return result["access_token"]

def _flush_cache(cache: "msal.SerializableTokenCache") -> None:  # type: ignore[name-defined]
    if cache.has_state_changed:
        AUTH_DIR.mkdir(parents=True, exist_ok=True)
        TOKEN_CACHE_PATH.write_text(cache.serialize(), encoding="utf-8")

# ── REST helpers ───────────────────────────────────────────────────────────────

def _rest_get(path: str, token: str, *, retry: int = 5) -> Any:
    """GET from the Power BI REST API with exponential back-off on 429."""
    url     = f"{API_BASE}{path}"
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    delay   = 1.0
    for attempt in range(retry):
        t0   = time.monotonic()
        resp = requests.get(url, headers=headers, timeout=30)
        dbg(f"GET {url}  →  HTTP {resp.status_code}  ({time.monotonic()-t0:.2f}s)")
        if resp.status_code == 200:
            payload = resp.json()
            count   = len(payload.get("value", [])) if isinstance(payload, dict) else "?"
            dbg(f"     {count} item(s) in response")
            return payload
        if resp.status_code == 429:
            wait = float(resp.headers.get("Retry-After", delay))
            print(yellow(f"  {ts()} 429 throttled — waiting {wait:.0f}s before retry {attempt+1}/{retry} …"))
            time.sleep(wait)
            delay = min(delay * 2, 60)
            continue
        if resp.status_code == 403:
            raise PermissionError(f"403 Forbidden: {url}")
        if resp.status_code == 404:
            raise FileNotFoundError(f"404 Not Found: {url}")
        resp.raise_for_status()
    raise RuntimeError(f"Exceeded retry limit for {url}")


def list_workspaces(token: str) -> list[dict]:
    data = _rest_get("/v1.0/myorg/groups?$top=1000", token)
    result = [{"id": w["id"], "name": w.get("displayName") or w.get("name", "Unknown")} for w in data.get("value", [])]
    dbg(f"list_workspaces → {len(result)} workspace(s)")
    return result

def list_datasets(token: str, workspace_id: str) -> list[dict]:
    data = _rest_get(f"/v1.0/myorg/groups/{workspace_id}/datasets", token)
    result = [{"id": d["id"], "name": d["name"]} for d in data.get("value", [])]
    dbg(f"list_datasets(ws={workspace_id}) → {len(result)} dataset(s)")
    return result

def get_datasources(token: str, workspace_id: str, dataset_id: str) -> list[dict]:
    dbg(f"get_datasources: ws={workspace_id}  ds={dataset_id}")
    try:
        data = _rest_get(
            f"/v1.0/myorg/groups/{workspace_id}/datasets/{dataset_id}/datasources",
            token,
        )
    except PermissionError as exc:
        print(yellow(f"      ⚠  {exc} — skipping datasource scan for this dataset"))
        return []
    except FileNotFoundError as exc:
        print(yellow(f"      ⚠  {exc} — dataset not found, skipping"))
        return []
    results = data.get("value", [])
    if _DEBUG:
        for src in results:
            dbg(f"     datasource: type={src.get('datasourceType')!r}  "
                f"conn={src.get('connectionDetails')}")
    return results

def _source_format_from_path(path: str) -> str:
    lower = path.lower().split("?")[0]  # strip query string before checking extension
    if lower.endswith(".xlsx") or lower.endswith(".xls") or lower.endswith(".xlsm"):
        return "xlsx"
    if lower.endswith(".csv"):
        return "csv"
    return "unknown"

_UPN_HEADER_HINTS = {
    "email", "upn", "userprincipalname", "user", "loginname", "login",
    "username", "samaccountname", "mail", "account",
    # broader enterprise/healthcare terms
    "emailaddress", "useraccess", "aduser", "adaccount", "networkid",
    "networklogin", "domainuser", "employeeemail", "workplaceemail",
    # FH-specific
    "fhemail", "fhmail",
}

_RLS_FILENAME_HINTS = {"user access", "useraccess", "user list", "userlist",
                       "security", "rls", "permission", "access control"}

def _find_header_row(ws, max_search: int = 25) -> tuple[list[str], int]:
    """
    Scan up to max_search rows to find the first row that looks like
    real column headers (≥2 non-empty short cells, no long prose sentences).
    Returns (headers, row_index_1based).
    """
    for i, row in enumerate(ws.iter_rows(max_row=max_search, values_only=True), start=1):
        cells = [str(v).strip() for v in row if v is not None and str(v).strip()]
        if len(cells) < 2:
            continue
        # Reject rows that look like prose (any single cell > 60 chars)
        if any(len(c) > 60 for c in cells):
            continue
        return cells, i
    return [], 1


def _sniff_upn_column(path: str, fmt: str, sheet: str | None) -> tuple[str | None, list[str]]:
    """
    Open a local/UNC xlsx/csv file and return (upn_column, all_headers).
    Handles multi-row preambles by scanning for the real header row.
    """
    if not path or not os.path.isfile(path):
        return None, []
    try:
        if fmt == "xlsx" and _OPENPYXL_OK:
            wb = _openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws_names = [sheet] if sheet and sheet in wb.sheetnames else list(wb.sheetnames)
            for ws_name in ws_names:
                ws = wb[ws_name]
                headers, hdr_row = _find_header_row(ws)
                if not headers:
                    continue
                match = _rank_upn_headers_by_name(headers)
                if not match:
                    # Sample data rows below the header for @ values
                    match = _rank_upn_headers_by_sample(headers, ws, hdr_row)
                wb.close()
                return match, headers
            wb.close()

        elif fmt == "csv":
            with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
                reader = _csv_mod.reader(f)
                headers = next(reader, [])
                sample_rows = [next(reader, []) for _ in range(5)]
            return _rank_upn_headers_from_samples(headers, sample_rows), headers

    except Exception as e:
        dbg(f"_sniff_upn_column: could not read {path!r}: {e}")
    return None, []


def _rank_upn_headers_by_name(headers: list[str]) -> str | None:
    """Match column names against UPN hint keywords."""
    for h in headers:
        if h.lower().replace(" ", "").replace("_", "") in _UPN_HEADER_HINTS:
            return h
    return None


def _rank_upn_headers_by_sample(headers: list[str], ws, hdr_row: int) -> str | None:
    """Sample rows below the header row looking for @ email values."""
    sample: dict[int, list[str]] = {i: [] for i in range(len(headers))}
    for row in ws.iter_rows(min_row=hdr_row + 1, max_row=hdr_row + 10, values_only=True):
        for i, val in enumerate(row):
            if val and i < len(headers):
                sample[i].append(str(val))
    for i, vals in sample.items():
        if any("@" in v for v in vals):
            return headers[i] if i < len(headers) else None
    return None


def _rank_upn_headers_from_samples(headers: list[str], rows: list[list[str]]) -> str | None:
    for h in headers:
        if h.lower().replace(" ", "").replace("_", "") in _UPN_HEADER_HINTS:
            return h
    for i, h in enumerate(headers):
        if any("@" in (row[i] if i < len(row) else "") for row in rows):
            return h
    return None


_DERIVED_TABLE_RE = re.compile(r'^\s*let\s+\w[\w\s]*=\s*#"', re.MULTILINE)

def _sniff_upn_from_m_expression(m_code: str) -> str | None:
    """
    Parse a Power Query M expression to identify the UPN column name.
    Returns None (without trying) for derived tables that reference another
    model table via  Source = #"Table Name"  — those inherit columns from
    the parent and are too idiosyncratic to resolve generically.
    """
    if not m_code:
        return None

    # Bail out early if this is a derived/downstream table referencing another
    # model table rather than a direct file or DB source
    if _DERIVED_TABLE_RE.search(m_code):
        dbg("  M-parse: derived table reference — skipping UPN sniff")
        return None

    # ── 1. Parse Table.RenameColumns ─────────────────────────────────────────
    rename_map: dict[str, str] = {}
    for old, new in re.findall(r'\{\s*"([^"]+)"\s*,\s*"([^"]+)"\s*\}', m_code):
        rename_map[old] = new

    for old, new in rename_map.items():
        if new.lower().replace(" ", "").replace("_", "") in _UPN_HEADER_HINTS:
            dbg(f"  M-parse UPN: rename {old!r} → {new!r} matches hint")
            return new
        if old.lower().replace(" ", "").replace("_", "") in _UPN_HEADER_HINTS:
            dbg(f"  M-parse UPN: original {old!r} renamed to {new!r}")
            return new

    # ── 2. Final column list from SelectColumns / ReorderColumns ─────────────
    select_match = re.search(
        r'Table\.(?:SelectColumns|ReorderColumns)\s*\([^,]+,\s*\{([^}]+)\}', m_code
    )
    if select_match:
        for col in re.findall(r'"([^"]+)"', select_match.group(1)):
            if col.lower().replace(" ", "").replace("_", "") in _UPN_HEADER_HINTS:
                dbg(f"  M-parse UPN: SelectColumns match {col!r}")
                return col

    # ── 3. Fallback: any quoted token matching a hint ─────────────────────────
    for token in re.findall(r'"([^"]+)"', m_code):
        if token.lower().replace(" ", "").replace("_", "") in _UPN_HEADER_HINTS:
            dbg(f"  M-parse UPN: fallback token match {token!r}")
            return token

    return None


def _file_source_rows_from_rest(
    workspace: dict,
    dataset:   dict,
    datasources: list[dict],
    timestamp: str,
) -> list[dict]:
    """Build Tier-1 manifest rows from REST datasource data."""
    dbg(f"_file_source_rows_from_rest: {len(datasources)} datasource(s) to inspect")
    rows = []
    for ds in datasources:
        ds_type = ds.get("datasourceType", "")
        conn    = ds.get("connectionDetails") or {}
        if ds_type not in FILE_SOURCE_TYPES:
            dbg(f"  skip  type={ds_type!r}  (not in FILE_SOURCE_TYPES)")
            continue
        path    = (
            conn.get("path")
            or conn.get("url")
            or conn.get("domain")
            or conn.get("server")
            or conn.get("connectionString")
            or ""
        )
        if not path:
            dbg(f"  skip  type={ds_type!r}  (empty connectionDetails — no path)")
            continue

        # Classify source type
        is_dynamics = (
            ds_type in ("Extension", "CommonDataService")
            or any(x in path.lower() for x in (".dynamics.com", ".crm", "commondata"))
        )
        is_ad = ds_type in ("ActiveDirectory", "Ldap") and not is_dynamics

        if is_dynamics:
            fmt = "dataverse"
            notes = "Dynamics 365 / Dataverse source — RLS role details require XMLA"
        elif is_ad:
            fmt = "active_directory"
            notes = "On-prem AD/LDAP source — validation requires LDAP query, not file open"
        else:
            fmt = _source_format_from_path(path)
            notes = "Tier 1 only — RLS role/table details require XMLA"

        dbg(f"  include  type={ds_type!r}  format={fmt!r}  path={path!r}")

        # Try to detect UPN column by reading the file directly (xlsx/csv only)
        upn_col = None
        file_headers: list[str] = []
        if fmt in ("xlsx", "csv"):
            upn_col, file_headers = _sniff_upn_column(path, fmt, None)
            if upn_col:
                dbg(f"  sniffed UPN column: {upn_col!r}")
            elif file_headers:
                # File was readable but no UPN column matched — report headers
                # so the user can see what's actually in the file
                fname = Path(path).name.lower()
                is_likely_rls = any(h in fname for h in _RLS_FILENAME_HINTS)
                col_preview = ", ".join(file_headers[:10])
                if is_likely_rls:
                    notes += f" — likely RLS file but UPN column not auto-detected; columns: [{col_preview}]"
                else:
                    notes += f" — columns found: [{col_preview}]"
            else:
                notes += " — file not accessible or empty"

        rows.append({
            "workspace_name":   workspace["name"],
            "workspace_id":     workspace["id"],
            "dataset_name":     dataset["name"],
            "dataset_id":       dataset["id"],
            "role_name":        None,
            "rls_table":        None,
            "dax_filter":       None,
            "upn_column":       upn_col,
            "source": {
                "path":   path,
                "file":   None if (is_ad or is_dynamics) else (Path(path.split("?")[0]).name or path),
                "format": fmt,
                "sheet":  None,
            },
            "discovery_method": "rest",
            "scan_timestamp":   timestamp,
            "notes":            notes,
        })
    dbg(f"_file_source_rows_from_rest → {len(rows)} row(s) produced")
    return rows

# ── XMLA scan (Tier 2) ─────────────────────────────────────────────────────────

def _extract_m_path(expression: str) -> tuple[str, str, str | None]:
    """
    Extract (path, format, sheet) from a Power Query M expression.

    Checks on-prem AD / LDAP patterns first, then file-based patterns.
    Returns ("", "unknown", None) when no recognised pattern is found.

    format values:
      "xlsx" | "csv"             — flat-file sources
      "active_directory" | "ldap" — on-prem AD / LDAP security-group sources
      "unknown"                  — source found but format unrecognised
    """
    dbg(f"_extract_m_path: expr snippet = {expression[:120]!r}")

    # ── on-prem AD / LDAP (must be checked before generic File.Contents) ─────
    for pattern, fmt in _AD_M_PATTERNS:
        m = pattern.search(expression)
        if m:
            dbg(f"  AD/LDAP match: pattern={pattern.pattern!r}  path={m.group(1)!r}  fmt={fmt!r}")
            return m.group(1), fmt, None

    # ── file-based sources ────────────────────────────────────────────────────
    for pattern in _M_PATH_PATTERNS:
        m = pattern.search(expression)
        if m:
            path   = m.group(1)
            fmt    = _source_format_from_path(path)
            # check for Excel.Workbook to be sure
            if "Excel.Workbook" in expression or path.lower().split("?")[0].endswith((".xlsx", ".xls", ".xlsm")):
                fmt = "xlsx"
            elif "Csv.Document" in expression or path.lower().endswith(".csv"):
                fmt = "csv"
            sheet_m = _M_SHEET_PATTERN.search(expression)
            sheet   = sheet_m.group(1) if sheet_m else None
            dbg(f"  file match: pattern={pattern.pattern[:60]!r}  path={path!r}  fmt={fmt!r}  sheet={sheet!r}")
            return path, fmt, sheet
    dbg("  no pattern matched — returning empty")
    return "", "unknown", None

def _xmla_rows(
    workspace:    dict,
    dataset:      dict,
    token:        str,
    timestamp:    str,
) -> list[dict]:
    """
    Tier 2: connect via XMLA, walk TOM to extract RLS roles + M expressions.
    Returns [] if pyadomd is unavailable or the connection fails.
    """
    if not _XMLA_LIB_OK:
        dbg("_xmla_rows: pyadomd not available — skipping Tier 2")
        return []

    from pyadomd import Pyadomd  # type: ignore[import]

    conn_str = (
        f"Provider=MSOLAP;"
        f"Data Source=powerbi://api.powerbi.com/v1.0/myorg/{workspace['name']};"
        f"Initial Catalog={dataset['name']};"
        f"User ID=;"
        f"Password=<token>;"   # password logged as placeholder for security
    )
    dbg(f"XMLA connecting: {conn_str}")
    # Rebuild with real token for actual connection
    conn_str = conn_str.replace("Password=<token>;", f"Password={token};")

    rows: list[dict] = []
    try:
        with Pyadomd(conn_str) as conn:
            dbg("XMLA connected")

            # ── 1a. Fetch roles (ID → name) ───────────────────────────────────
            dbg("XMLA: querying TMSCHEMA_ROLES …")
            roles: dict[str, str] = {}
            with conn.cursor().execute(
                "SELECT [ID], [NAME] FROM $SYSTEM.TMSCHEMA_ROLES"
            ) as cur:
                for row in cur.fetchall():
                    roles[row[0]] = row[1]
            dbg(f"  {len(roles)} role(s) in model")

            # ── 1b. Fetch table permissions (no JOIN — not supported in DMX) ──
            dbg("XMLA: querying TMSCHEMA_TABLE_PERMISSIONS …")
            role_rows: list[tuple[str, str, str, str]] = []
            with conn.cursor().execute(
                "SELECT [ROLE_ID], [TABLE_ID], [FILTER_EXPRESSION] "
                "FROM $SYSTEM.TMSCHEMA_TABLE_PERMISSIONS"
            ) as cur:
                for row in cur.fetchall():
                    role_id, table_id, dax_filter = row
                    role_name = roles.get(role_id, role_id)
                    has_upn = "USERPRINCIPALNAME" in (dax_filter or "").upper()
                    dbg(f"  role={role_name!r}  table_id={table_id!r}  "
                        f"has_UPN={has_upn}  filter={str(dax_filter)[:80]!r}")
                    if not has_upn:
                        continue
                    upn_m = _DAX_UPN_PATTERN.search(dax_filter or "")
                    upn_col = upn_m.group(1) if upn_m else "(complex — check DAX manually)"
                    dbg(f"  → UPN role kept: role={role_name!r}  upn_column={upn_col!r}")
                    role_rows.append((role_name, table_id, dax_filter, upn_col))

            dbg(f"  {len(role_rows)} UPN role(s) found")
            if not role_rows:
                dbg("No USERPRINCIPALNAME roles — skipping table/partition queries")
                return []

            # ── 2. Fetch table names keyed by ID ──────────────────────────────
            table_query = "SELECT [ID], [NAME] FROM $SYSTEM.TMSCHEMA_TABLES"
            dbg("XMLA: querying TMSCHEMA_TABLES …")
            table_names: dict[str, str] = {}
            with conn.cursor().execute(table_query) as cur:
                for row in cur.fetchall():
                    table_names[row[0]] = row[1]
            dbg(f"  {len(table_names)} table(s) in model")

            # ── 3. Fetch M partition expressions ──────────────────────────────
            part_query = """
                SELECT
                    PARTITIONS.[TABLE_ID],
                    PARTITIONS.[NAME],
                    PARTITIONS.[QUERY_DEFINITION]
                FROM $SYSTEM.TMSCHEMA_PARTITIONS
                WHERE PARTITIONS.[SOURCE_TYPE] = 2   -- M partitions
            """
            dbg("XMLA: querying TMSCHEMA_PARTITIONS (SOURCE_TYPE=2, M only) …")
            m_by_table: dict[str, list[tuple[str, str]]] = {}
            with conn.cursor().execute(part_query) as cur:
                for row in cur.fetchall():
                    tid, pname, expr = row
                    m_by_table.setdefault(tid, []).append((pname, expr or ""))
            dbg(f"  {sum(len(v) for v in m_by_table.values())} M partition(s) across "
                f"{len(m_by_table)} table(s)")

            # ── 4. Build manifest rows ─────────────────────────────────────────
            for role_name, table_id, dax_filter, upn_col in role_rows:
                rls_table  = table_names.get(table_id, table_id)
                partitions = m_by_table.get(table_id, [])
                dbg(f"Building row: role={role_name!r}  table={rls_table!r}  "
                    f"partitions={len(partitions)}")

                if not partitions:
                    print(f"      {dim('↳')}  {yellow('⚑')}  {dim(f'role={role_name!r}  table={rls_table!r} — embedded (no M partition)')}")
                    rows.append({
                        "workspace_name":   workspace["name"],
                        "workspace_id":     workspace["id"],
                        "dataset_name":     dataset["name"],
                        "dataset_id":       dataset["id"],
                        "role_name":        role_name,
                        "rls_table":        rls_table,
                        "dax_filter":       dax_filter,
                        "upn_column":       upn_col,
                        "source": {
                            "path":   None,
                            "file":   None,
                            "format": "embedded",
                            "sheet":  None,
                        },
                        "discovery_method": "xmla",
                        "scan_timestamp":   timestamp,
                        "notes":            "No M partition found — table may be embedded in the model",
                    })
                    continue

                for pname, expr in partitions:
                    dbg(f"  partition={pname!r}  expr_len={len(expr)}")
                    path, fmt, sheet = _extract_m_path(expr)

                    # Enrich UPN column from M expression when DAX regex wasn't specific
                    resolved_upn = upn_col
                    if resolved_upn in ("(complex — check DAX manually)", None):
                        m_upn = _sniff_upn_from_m_expression(expr)
                        if m_upn:
                            resolved_upn = m_upn
                            dbg(f"  M-parse resolved UPN column: {m_upn!r}")

                    if not path:
                        # Non-file source (e.g. Dataverse/CRM) — still record if we found a UPN column
                        if resolved_upn and resolved_upn not in ("(complex — check DAX manually)",):
                            rows.append({
                                "workspace_name":   workspace["name"],
                                "workspace_id":     workspace["id"],
                                "dataset_name":     dataset["name"],
                                "dataset_id":       dataset["id"],
                                "role_name":        role_name,
                                "rls_table":        rls_table,
                                "dax_filter":       dax_filter,
                                "upn_column":       resolved_upn,
                                "source": {
                                    "path":   None,
                                    "file":   None,
                                    "format": "dataverse" if "CommonDataService" in expr or "Dataverse" in expr else "non-file",
                                    "sheet":  None,
                                },
                                "discovery_method": "xmla",
                                "scan_timestamp":   timestamp,
                                "notes":            f"Non-file source (partition: {pname})",
                            })
                        continue

                    rows.append({
                        "workspace_name":   workspace["name"],
                        "workspace_id":     workspace["id"],
                        "dataset_name":     dataset["name"],
                        "dataset_id":       dataset["id"],
                        "role_name":        role_name,
                        "rls_table":        rls_table,
                        "dax_filter":       dax_filter,
                        "upn_column":       resolved_upn,
                        "source": {
                            "path":   path,
                            "file":   Path(path.split("?")[0]).name or path,
                            "format": fmt,
                            "sheet":  sheet,
                        },
                        "discovery_method": "xmla",
                        "scan_timestamp":   timestamp,
                        "notes":            "",
                    })

    except Exception as exc:  # noqa: BLE001
        import traceback as _tb
        dbg(f"XMLA failed for {dataset['name']}: {exc}\n" + _tb.format_exc())
        print(yellow(f"  ⚠  XMLA unavailable for {dataset['name'][:50]} — using REST fallback"), flush=True)
        return []

    dbg(f"_xmla_rows → {len(rows)} row(s) for dataset {dataset['name']!r}")
    return rows

# ── deduplication ──────────────────────────────────────────────────────────────

def _dedup(rows: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    out:  list[dict] = []
    for r in rows:
        key = (r["dataset_id"], r["role_name"], (r["source"] or {}).get("path") or "")
        if key not in seen:
            seen.add(key)
            out.append(r)
    return sorted(out, key=lambda r: (
        r["workspace_name"].lower(),
        r["dataset_name"].lower(),
        (r["role_name"] or "").lower(),
    ))

# ── manifest write ─────────────────────────────────────────────────────────────

def write_manifest(rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        yaml.dump(
            rows,
            f,
            allow_unicode    = True,
            default_flow_style = False,
            sort_keys        = False,
        )

# ── summary table ──────────────────────────────────────────────────────────────

def _print_results(rows: list[dict], verbose: bool) -> None:
    limit   = len(rows) if verbose else min(50, len(rows))
    display = rows[:limit]
    width   = max((len(r["dataset_name"]) for r in rows), default=20) + 2

    print(f"\n  {bold('Results')}{dim(f' — {len(rows)} entr' + ('y' if len(rows) == 1 else 'ies'))}")
    header = (
        f"  {'Dataset':<{width}}  {'Role':<25}  {'Column':<20}  {'Format':<8}  Path"
    )
    print(dim(header))
    print(dim("  " + "─" * (len(header) - 2)))

    for r in display:
        source  = r["source"] or {}
        dataset = r["dataset_name"][:width - 1].ljust(width)
        role    = (r["role_name"] or "")[:24].ljust(25)
        col     = (r["upn_column"] or "")[:19].ljust(20)
        fmt     = (source.get("format") or "")[:7].ljust(8)
        path    = source.get("path") or ""
        method  = r.get("discovery_method", "")
        method_tag = dim(f"[{method}]") if method else ""
        print(f"  {cyan(dataset)}  {role}  {green(col)}  {fmt}  {dim(path)} {method_tag}")

    if not verbose and len(rows) > 50:
        print(dim(f"\n  … {len(rows) - 50} more entries — rerun with --verbose to see all"))

# ── scan one workspace ─────────────────────────────────────────────────────────

def scan_workspace(
    workspace:      dict,
    token:          str,
    no_xmla:        bool,
    timestamp:      str,
    dataset_filter: str | None = None,
) -> list[dict]:
    all_datasets = list_datasets(token, workspace["id"])

    # Apply dataset filter (exact match first, then fuzzy substring)
    if dataset_filter:
        fl = dataset_filter.lower()
        datasets = [d for d in all_datasets if d["name"].lower() == fl]
        if not datasets:
            datasets = [d for d in all_datasets if fl in d["name"].lower()]
        if not datasets:
            print(yellow(f"  No dataset matching {dataset_filter!r} in {workspace['name']}"))
            return []
    else:
        datasets = all_datasets

    total = len(datasets)
    ws_id = workspace["id"]
    tier  = 'Tier 2 (XMLA)' if (_XMLA_LIB_OK and not no_xmla) else 'Tier 1 (REST only)'
    print(f"\n  {ts()} {bold(cyan(workspace['name']))}  {dim(f'({ws_id})')}")
    print(f"  {ts()} {total} dataset(s) to scan  •  {tier}")

    all_rows: list[dict] = []

    for idx, ds in enumerate(datasets, 1):
        counter = dim(f"[{idx:>{len(str(total))}}/{total}]")
        start   = time.monotonic()
        rows: list[dict] = []

        print(f"    {counter}  {ds['name'][:50]}", end="", flush=True)
        dbg(f"  dataset id={ds['id']!r}")

        # ── Tier 2: XMLA ──────────────────────────────────────────────────────
        if not no_xmla and _XMLA_LIB_OK:
            rows = _xmla_rows(workspace, ds, token, timestamp)
            if rows:
                dbg(f"  XMLA produced {len(rows)} row(s) — skipping REST")

        # ── Tier 1: REST (fallback or supplement) ─────────────────────────────
        if not rows:
            if not no_xmla and _XMLA_LIB_OK:
                print(dim(" → fallback REST …"), end="", flush=True)
            else:
                print(dim("  [REST] …"), end="", flush=True)
            datasources = get_datasources(token, workspace["id"], ds["id"])
            tier1       = _file_source_rows_from_rest(workspace, ds, datasources, timestamp)
            rows.extend(tier1)

        count = len(rows)
        el    = elapsed(start)
        if count:
            method    = rows[0]["discovery_method"] if rows else "?"
            formats   = ", ".join(sorted({r["source"]["format"] for r in rows if r["source"]}))
            upn_cols  = ", ".join(sorted({
                r["upn_column"] for r in rows
                if r["upn_column"] not in (None, "")
            })) or dim("(unknown)")
            print(f"\r    {counter}  {green('✔')}  {ds['name'][:50]:<52}"
                  f"  {dim(f'[{method}]')}  {count} row(s)  "
                  f"fmt={dim(formats)}  col={cyan(upn_cols)}  {el}")
        else:
            print(f"\r    {counter}  {dim('–')}  {dim(ds['name'][:50]):<52}"
                  f"  {dim('no RLS file sources')}  {el}")

        all_rows.extend(rows)

    print(f"\n  {ts()} {bold(workspace['name'])} complete — "
          f"{green(str(len(all_rows)))} total row(s) found")
    return all_rows

# ── main ───────────────────────────────────────────────────────────────────────

def main() -> None:
    global _DEBUG

    parser = argparse.ArgumentParser(
        prog        = "discover_rls_sources.py",
        description = "Scan Power BI workspaces for RLS source files and write rls_sources_manifest.yaml",
    )
    parser.add_argument("--all",       action="store_true", help="Scan all workspaces without prompting")
    parser.add_argument("--workspace", metavar="NAME",      help="Scan a specific workspace by name (skips picker)")
    parser.add_argument("--dataset",   metavar="NAME",      help="Scan only a specific dataset/model by name (skips dataset picker)")
    parser.add_argument("--output",    metavar="DIR",       help=f"Write manifest to this directory (default: rls_discovery/)")
    parser.add_argument("--no-xmla",   action="store_true", help="Tier 1 (REST) only — skip XMLA even if pyadomd is installed")
    parser.add_argument("--verbose",   action="store_true", help="Print all result rows (default: first 50)")
    parser.add_argument("--debug",     action="store_true", help="Print diagnostic output for every API call, M expression, and XMLA query")
    args = parser.parse_args()

    _DEBUG = args.debug

    output_path = (
        Path(args.output) / "rls_sources_manifest.yaml"
        if args.output
        else MANIFEST_PATH
    )

    print()
    print(f"  {bold('Power BI RLS Source Discovery')} ({yellow('Tier 1 REST') if not _XMLA_LIB_OK else green('Tier 2 XMLA')})")
    if _DEBUG:
        print(yellow("  DEBUG MODE ON — verbose diagnostic output enabled"))
    print()

    # ── auth ──────────────────────────────────────────────────────────────────
    t0 = time.monotonic()
    print(f"  {ts()} Checking token cache …", end=" ", flush=True)
    cache_exists = TOKEN_CACHE_PATH.exists()
    print(dim("found.") if cache_exists else yellow("not found — device flow required."))

    token = get_access_token()
    print(f"  {ts()} Authenticated.  {elapsed(t0)}")

    if _XMLA_DLL_MISSING and not args.no_xmla:
        print(yellow(f"  ⚠  XMLA DLLs not found — falling back to Tier 1 (REST only)"))
        for m in _XMLA_DLL_MISSING:
            print(dim(f"       missing: {m}"))
        print(dim("       Install SSMS 21 or adjust _DLL_PATHS in the script."))

    xmla_note = (
        dim("XMLA (Tier 2) available")
        if (_XMLA_LIB_OK and not args.no_xmla)
        else yellow("Tier 1 only") + dim(" (DLLs missing — see above)")
        if _XMLA_DLL_MISSING
        else yellow("Tier 1 only") + dim(" (install pyadomd for full XMLA scanning)")
        if not _XMLA_LIB_OK
        else dim("Tier 1 only (--no-xmla)")
    )
    print(f"  {ts()} Scan mode: {xmla_note}")
    print()

    # ── workspace selection ───────────────────────────────────────────────────
    workspaces = list_workspaces(token)

    if args.workspace:
        name_lower = args.workspace.lower()
        match = next((w for w in workspaces if w["name"].lower() == name_lower), None)
        if not match:
            # fuzzy fallback
            match = next((w for w in workspaces if name_lower in w["name"].lower()), None)
        if not match:
            print(red(f"  Workspace not found: {args.workspace}"))
            print(dim(f"  Available workspaces: {', '.join(w['name'] for w in workspaces[:10])}"))
            sys.exit(1)
        targets = [match]

    elif args.all:
        targets = workspaces

    else:
        print(f"  {bold('Scan scope')}")
        print(f"    {dim('[1]')}  All workspaces  {dim(f'({len(workspaces)} total — may be slow)')}")
        print(f"    {dim('[2]')}  Pick a workspace")
        choice = input(f"  {dim('> ')}").strip()

        if choice == "1":
            targets = workspaces
        else:
            targets = [pick_one(workspaces, "Workspaces")]

    print()

    # ── optional dataset filter ───────────────────────────────────────────────
    dataset_filter: str | None = None

    if args.dataset:
        dataset_filter = args.dataset
    elif len(targets) == 1:
        # Single workspace selected — offer dataset picker
        print(f"  {bold('Dataset scope')}")
        print(f"    {dim('[1]')}  All datasets in workspace")
        print(f"    {dim('[2]')}  Pick a specific dataset")
        ds_choice = input(f"  {dim('> ')}").strip()
        if ds_choice == "2":
            all_datasets = list_datasets(token, targets[0]["id"])
            picked_ds = pick_one(all_datasets, "Datasets")
            dataset_filter = picked_ds["name"]
            print()

    # ── scan ──────────────────────────────────────────────────────────────────
    timestamp = now_iso()
    all_rows: list[dict] = []

    for ws in targets:
        rows = scan_workspace(
            ws, token,
            no_xmla=args.no_xmla,
            timestamp=timestamp,
            dataset_filter=dataset_filter,
        )
        all_rows.extend(rows)

    all_rows = _dedup(all_rows)

    # ── output ────────────────────────────────────────────────────────────────
    datasets_with_results = len({r["dataset_name"] for r in all_rows})
    roles_found           = len({(r["dataset_id"], r["role_name"]) for r in all_rows})
    file_sources          = len([r for r in all_rows if (r["source"] or {}).get("path")])
    needs_xmla = len([r for r in all_rows if r.get("role_name") is None])
    ad_sources            = len([r for r in all_rows if (r["source"] or {}).get("format") in ("active_directory", "ldap")])

    print()
    print(f"  {bold('Done.')}")
    print(f"    {cyan(str(datasets_with_results))} dataset(s) with RLS results")
    print(f"    {cyan(str(roles_found))} RLS role(s) using USERPRINCIPALNAME()")
    print(f"    {cyan(str(file_sources))} file source(s) identified")
    if ad_sources:
        print(f"    {cyan(str(ad_sources))} on-prem AD / LDAP source(s)  {dim('(no file — see notes)')}")
    if needs_xmla:
        xmla_note = '(role/table/filter unknown — REST tier only)'
        print(f"    {yellow(str(needs_xmla))} row(s) missing role detail  "
              f"{dim(xmla_note)}")
    if not _XMLA_LIB_OK and not args.no_xmla:
        print(dim("\n  Tip: install pyadomd on Windows to get upn_column values without manual lookup."))
        print(dim("       pip install pyadomd"))

    write_manifest(all_rows, output_path)
    print(f"\n  {green('→')} {bold(str(output_path))}  {dim(f'({len(all_rows)} entr' + ('y' if len(all_rows)==1 else 'ies') + ')')}")
    if all_rows:
        _print_results(all_rows, verbose=args.verbose)
    else:
        print(dim("\n  No RLS file sources found in the scanned workspace(s)."))
        print(dim("  Possible reasons: no USERPRINCIPALNAME() filters, no file-backed datasources,"))
        print(dim("  or insufficient permissions (403) on some datasets."))
        if not _XMLA_LIB_OK:
            print(dim("  Note: without XMLA only REST datasources are visible — SQL-backed RLS tables"))
            print(dim("  will not appear here even if they use USERPRINCIPALNAME()."))

    print()


if __name__ == "__main__":
    main()
